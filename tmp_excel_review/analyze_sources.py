import json
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DOCS = ROOT / "source_docs"


def pdf_data(path: Path) -> dict:
    pages = []
    with pdfplumber.open(path) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            pages.append({
                "page": index,
                "width": page.width,
                "height": page.height,
                "text": page.extract_text(x_tolerance=2, y_tolerance=3) or "",
                "tables": page.extract_tables(),
            })
    return {"file": path.name, "page_count": len(pages), "pages": pages}


def value_preview(ws, max_rows=120, max_cols=60):
    rows = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, max_rows),
        min_col=1,
        max_col=min(ws.max_column, max_cols),
    ):
        values = [cell.value for cell in row]
        if any(value not in (None, "") for value in values):
            rows.append({"row": row[0].row, "values": values})
    return rows


def workbook_data(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        formulas = []
        hyperlinks = []
        comments = []
        styled_cells = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
                if cell.hyperlink:
                    hyperlinks.append({"cell": cell.coordinate, "target": cell.hyperlink.target})
                if cell.comment:
                    comments.append({"cell": cell.coordinate, "author": cell.comment.author, "text": cell.comment.text})
                if cell.has_style:
                    styled_cells += 1

        validations = []
        for validation in ws.data_validations.dataValidation:
            validations.append({
                "ranges": str(validation.sqref),
                "type": validation.type,
                "formula1": validation.formula1,
                "formula2": validation.formula2,
            })

        conditional = []
        for area, rules in ws.conditional_formatting._cf_rules.items():
            conditional.append({
                "ranges": str(area.sqref),
                "rules": [
                    {
                        "type": rule.type,
                        "operator": rule.operator,
                        "formula": rule.formula,
                        "text": rule.text,
                    }
                    for rule in rules
                ],
            })

        tables = [
            {"name": table.name, "display_name": table.displayName, "ref": table.ref}
            for table in ws.tables.values()
        ]
        sheets.append({
            "name": ws.title,
            "dimensions": ws.calculate_dimension(),
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
            "hidden_rows": [index for index, dim in ws.row_dimensions.items() if dim.hidden],
            "hidden_columns": [index for index, dim in ws.column_dimensions.items() if dim.hidden],
            "tables": tables,
            "validations": validations,
            "conditional_formatting": conditional,
            "formulas": formulas,
            "hyperlinks": hyperlinks,
            "comments": comments,
            "styled_cells": styled_cells,
            "rows": value_preview(ws),
        })
    return {"file": path.name, "sheet_count": len(sheets), "sheets": sheets}


payload = {
    "pdfs": [pdf_data(path) for path in sorted(DOCS.glob("*.pdf"))],
    "workbook": workbook_data(ROOT / "Janos Quinta y Pilar Hotel.xlsx"),
}
(ROOT / "sources_analysis.json").write_text(
    json.dumps(payload, ensure_ascii=False, indent=2, default=str),
    encoding="utf-8",
)

print(json.dumps({
    "pdfs": [{"file": item["file"], "pages": item["page_count"]} for item in payload["pdfs"]],
    "workbook": {
        "file": payload["workbook"]["file"],
        "sheets": [
            {"name": sheet["name"], "dimensions": sheet["dimensions"], "rows_with_values": len(sheet["rows"])}
            for sheet in payload["workbook"]["sheets"]
        ],
    },
}, ensure_ascii=True, indent=2))
